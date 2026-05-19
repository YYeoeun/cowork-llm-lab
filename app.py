import json
import os
import shutil
import subprocess
import tempfile
import urllib.error
import urllib.request
from collections import defaultdict
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html
from xlsx_merge import merge_workbooks as xml_merge_workbooks

LIBREOFFICE_BIN = shutil.which("soffice") or shutil.which("libreoffice")
PDFTOPPM_BIN = shutil.which("pdftoppm")
HAS_LIBREOFFICE = bool(LIBREOFFICE_BIN and PDFTOPPM_BIN)

OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://192.168.100.90:11434")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "gemma4:26b")


def ollama_chat(messages, format_json=False, model=None, timeout=180):
    payload = {
        "model": model or OLLAMA_MODEL,
        "messages": messages,
        "stream": False,
    }
    if format_json:
        payload["format"] = "json"
    req = urllib.request.Request(
        f"{OLLAMA_HOST}/api/chat",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.URLError as e:
        raise RuntimeError(f"Ollama 호출 실패: {e}") from e
    return data["message"]["content"]


def sense_sheet(xlsx_bytes, sheet_name, max_rows=20, max_cols=20):
    """Compact textual snapshot of a sheet for LLM analysis.
    Includes merged-cell ranges and per-cell bold/fill hints to expose structure
    that pure values would miss."""
    wb = load_workbook(BytesIO(xlsx_bytes))
    if sheet_name not in wb.sheetnames:
        return ""
    ws = wb[sheet_name]
    lines = [f"Sheet: {sheet_name}"]

    merges = [str(r) for r in ws.merged_cells.ranges]
    if merges:
        lines.append("Merged cells: " + ", ".join(merges))

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    scan_rows = min(max_rows, total_rows)
    scan_cols = min(max_cols, total_cols)
    lines.append(
        f"Shape: {total_rows} rows × {total_cols} cols. Showing top {scan_rows} × {scan_cols}."
    )
    header_letters = [get_column_letter(c) for c in range(1, scan_cols + 1)]
    lines.append("    | " + " | ".join(header_letters))

    for r in range(1, scan_rows + 1):
        cells = []
        for c in range(1, scan_cols + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                cells.append("·")
                continue
            text = str(v)
            if len(text) > 24:
                text = text[:24] + "…"
            hints = []
            if cell.font and cell.font.bold:
                hints.append("B")
            fg = (
                cell.fill.fgColor.rgb
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb
                else None
            )
            if fg and fg not in ("00000000", "FFFFFFFF"):
                hints.append("bg")
            cells.append(text + (f"[{''.join(hints)}]" if hints else ""))
        lines.append(f"R{r:>3} | " + " | ".join(cells))
    return "\n".join(lines)


def sense_workbook(xlsx_bytes, file_name):
    wb = load_workbook(BytesIO(xlsx_bytes))
    blocks = [f"=== File: {file_name} ==="]
    for sn in wb.sheetnames:
        blocks.append(sense_sheet(xlsx_bytes, sn))
    return "\n\n".join(blocks)


def ollama_chat_json(messages, retries=1, **kwargs):
    last_err = None
    for _ in range(retries + 1):
        text = ollama_chat(messages, format_json=True, **kwargs)
        try:
            return json.loads(text)
        except json.JSONDecodeError as e:
            last_err = e
            messages = messages + [
                {"role": "assistant", "content": text},
                {
                    "role": "user",
                    "content": "JSON 파싱 실패. 유효한 JSON만 다시 출력해주세요.",
                },
            ]
    raise RuntimeError(f"Ollama JSON 파싱 실패: {last_err}")


def robust_json_parse(text):
    """Try several strategies to extract a JSON object from an LLM response.
    Returns the parsed dict, or None if all strategies fail."""
    if not text:
        return None
    candidates = []
    candidates.append(text)
    stripped = text.strip()
    candidates.append(stripped)

    # Strip markdown code fences (```json ... ``` or ``` ... ```)
    if stripped.startswith("```"):
        lines = stripped.split("\n")
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        candidates.append("\n".join(lines))

    # Outermost {...} block
    first = stripped.find("{")
    last = stripped.rfind("}")
    if first != -1 and last > first:
        candidates.append(stripped[first : last + 1])

    for c in candidates:
        try:
            obj = json.loads(c)
            if isinstance(obj, dict):
                return obj
        except json.JSONDecodeError:
            continue
    return None


def llm_get_json_with_retry(messages, max_attempts=3):
    """Call LLM with format_json, parse robustly, retry with error feedback.
    Returns (parsed_dict_or_None, last_raw, error_string_or_None, attempts_log)."""
    msgs = list(messages)
    last_raw = None
    last_error = None
    attempts_log = []
    for attempt in range(1, max_attempts + 1):
        try:
            raw = ollama_chat(msgs, format_json=True)
        except Exception as e:
            last_error = f"{type(e).__name__}: {e}"
            attempts_log.append({"attempt": attempt, "raw": None, "error": last_error})
            break
        last_raw = raw
        parsed = robust_json_parse(raw)
        if parsed is not None and parsed:
            attempts_log.append({"attempt": attempt, "raw": raw, "error": None})
            return parsed, raw, None, attempts_log
        last_error = "응답이 유효한 JSON 객체가 아닙니다."
        attempts_log.append({"attempt": attempt, "raw": raw, "error": last_error})
        msgs = msgs + [
            {"role": "assistant", "content": raw or ""},
            {
                "role": "user",
                "content": (
                    "이전 응답이 유효한 JSON 객체가 아닙니다. "
                    "마크다운 펜스, 설명, thinking 텍스트 없이 순수 JSON 객체만 다시 출력해주세요. "
                    "응답은 반드시 '{' 로 시작해서 '}' 로 끝나야 하며 그 안에 'groups' 키가 있어야 합니다."
                ),
            },
        ]
    return None, last_raw, last_error, attempts_log


PLAN_SCHEMA_DOC = """\
Plan JSON schema (template + union merge):
{
  "groups": [
    {
      "label": "<group name in Korean>",
      "sheets": [
        {
          "file": "<exact filename>",
          "sheet": "<exact sheet name>",
          "header_row_count": <int>,
          "key_column_count": <int>,
          "skip_keywords": ["<text in first column to mark non-data rows>"]
        }
      ]
    }
  ],
  "ambiguities": ["<Korean question>", ...]
}

Rules:
- For each sheet, identify just 2 numbers:
  · header_row_count: how many rows from the TOP are header rows (NOT data).
    Single-row header → 1. Two-row header (super-header + sub-header merged) → 2.
  · key_column_count: how many columns from the LEFT are key/identifier columns.
    Everything to the right of these columns is treated as a numeric measure and aggregated by SUM.
    E.g., for "비목분류 | 번호 | 비용명 | 계획예산 | 이월예산 | ..." → key_column_count = 3.
- skip_keywords: text in the FIRST column that marks a row as subtotal/total/aggregate (not raw data).
  Default common values: "소계", "합계", "총계", "내부흡수액", "외부유출액", "누계".
  These rows are skipped during key matching (their original cells stay in the template).
- The first sheet in each group acts as the LAYOUT TEMPLATE. Its rows, styles, merges, fonts are
  preserved verbatim in the summary. Other sources contribute values via SUM formulas.
- Keys are matched by tuple equality on the leftmost key_column_count cells.
  Rows from non-template sources whose key is NOT in the template are appended at the bottom.
- Group sheets together only if they share the same logical structure (same key columns + same
  measure columns in the same order).
- Only output JSON. No commentary, no markdown fences."""


ANALYZE_SYSTEM = (
    "You are a Korean-friendly data analyst that plans how to integrate multiple "
    "Excel sheets. Each sheet snapshot shows top rows with cell hints "
    "([B]=bold, [bg]=colored background). '·' means empty. Merged cell ranges "
    "are listed separately. Your job: read the snapshots, identify titles, "
    "headers, subtotal rows, and column roles. Then propose how to integrate.\n\n"
    + PLAN_SCHEMA_DOC
)


REFINE_SYSTEM = (
    "You are refining an Excel integration plan based on user instructions. "
    "Apply the user's correction and output the FULL updated plan as JSON. "
    "Keep unchanged parts intact.\n\n" + PLAN_SCHEMA_DOC
)


PER_FILE_SYSTEM = """\
You are analyzing a single Excel file's structure for later integration.
The snapshot shows top rows with cell hints ([B]=bold, [bg]=colored background).
'·' means empty. Merged cell ranges are listed separately.

Write a SHORT analysis note in Korean (plain prose, 5-10 lines max). For each sheet, cover:
- 한 줄 요약: 어떤 데이터로 보이는지
- 헤더 행 개수: 위에서부터 몇 행이 헤더인지 (1개행이면 1, super-header + sub-header이면 2)
- 항목(키) 열 개수: 왼쪽에서부터 몇 열이 식별자(예: 비목분류, 번호, 비용명) 인지.
  그 오른쪽 모든 열은 측정값으로 SUM 통합됩니다.
- 소계/합계 패턴: 첫 열에 어떤 키워드가 들어 있는 행이 합계/소계인지 (예: "소계", "내부흡수액")
- 통합 시 다른 파일과의 차이점이나 주의점 (있다면)

출력은 plain text 한국어만. JSON, 마크다운 헤더, 코드 펜스 금지."""


def analyze_per_file(sense_text):
    return ollama_chat(
        [
            {"role": "system", "content": PER_FILE_SYSTEM},
            {"role": "user", "content": sense_text},
        ]
    )


def analyze_initial(sense_blocks, notes_by_file=None):
    parts = []
    if notes_by_file:
        notes_section = "\n\n".join(
            f"--- {fn}에 대한 분석 노트 ---\n{note.strip()}"
            for fn, note in notes_by_file.items() if note and note.strip()
        )
        if notes_section:
            parts.append("각 파일에 대한 사전 분석 노트:\n\n" + notes_section)
    parts.append("Excel 시트 원본 스냅샷:\n\n" + "\n\n".join(sense_blocks))
    parts.append("위 노트와 스냅샷을 바탕으로 통합 plan JSON을 출력해주세요.")
    return ollama_chat_json(
        [
            {"role": "system", "content": ANALYZE_SYSTEM},
            {"role": "user", "content": "\n\n".join(parts)},
        ]
    )


DEFAULT_SKIP_KEYWORDS = ["소계", "합계", "총계", "내부흡수액", "외부유출액", "누계"]


def build_heuristic_plan(file_bytes_by_name):
    """Best-effort plan with header_row_count and key_column_count guesses."""
    by_signature = {}
    for fn, b in file_bytes_by_name.items():
        try:
            excel = pd.ExcelFile(BytesIO(b))
        except Exception:
            continue
        for sn in excel.sheet_names:
            raw = excel.parse(sn, header=None)
            if raw.empty:
                continue
            header_idx = detect_header_row(raw)
            header_n = header_idx + 1
            # Heuristic for key_column_count: count leading non-numeric columns
            wb_local = load_workbook(BytesIO(b), data_only=True)
            ws = wb_local[sn]
            max_col = ws.max_column or 0
            key_n = 0
            for c in range(1, max_col + 1):
                values_below = []
                for r in range(header_n + 1, min((ws.max_row or 0) + 1, header_n + 11)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        values_below.append(v)
                if not values_below:
                    break
                # If all are non-numeric-string-like, count as key
                if all(parse_numeric(v) is None for v in values_below):
                    key_n += 1
                else:
                    break
            if key_n < 1:
                key_n = 1
            spec = {
                "file": fn,
                "sheet": sn,
                "header_row_count": header_n,
                "key_column_count": key_n,
                "skip_keywords": list(DEFAULT_SKIP_KEYWORDS),
            }
            sig = (header_n, key_n, max_col)
            by_signature.setdefault(sig, []).append(spec)

    groups = []
    for idx, (sig, specs) in enumerate(by_signature.items(), 1):
        groups.append({
            "label": f"휴리스틱 그룹 {idx}",
            "sheets": specs,
        })
    return {
        "groups": groups,
        "ambiguities": [
            "LLM이 plan을 만들지 못해 휴리스틱으로 작성되었습니다. header_row_count / key_column_count를 챗봇으로 확인해주세요."
        ],
    }


def refine_plan(plan, user_msg, history):
    msgs = [{"role": "system", "content": REFINE_SYSTEM}]
    msgs.append(
        {
            "role": "user",
            "content": f"현재 plan:\n{json.dumps(plan, ensure_ascii=False, indent=2)}",
        }
    )
    msgs.append({"role": "assistant", "content": "확인했습니다."})
    for turn in history[-6:]:
        msgs.append({"role": turn["role"], "content": turn["content"]})
    msgs.append({"role": "user", "content": user_msg})
    return ollama_chat_json(msgs)

st.set_page_config(page_title="cowork-llm-lab", layout="wide")

st.title("cowork-llm-lab")
st.caption("LLM과 함께 협업하는 Streamlit 작업 도구")


def parse_numeric(value):
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_numeric_series(series):
    if pd.api.types.is_numeric_dtype(series):
        return True
    non_null = series.dropna()
    if non_null.empty:
        return False
    return all(parse_numeric(v) is not None for v in non_null)


def _format_cell(v):
    if pd.isna(v):
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        f = float(v)
        if f.is_integer():
            return f"{int(f):,}"
        return f"{f:,.2f}"
    return str(v)


def format_for_display(df):
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(_format_cell)
    return out


def detect_header_row(raw, max_scan=15):
    if raw.empty:
        return 0
    limit = min(max_scan, len(raw))
    densities = [raw.iloc[i].notna().sum() for i in range(limit)]
    peak = max(densities)
    if peak == 0:
        return 0
    threshold = max(2, int(peak * 0.7))

    best_idx = 0
    best_score = -1
    for i in range(limit):
        row = raw.iloc[i].dropna()
        if len(row) < threshold:
            continue
        if any(parse_numeric(v) is not None for v in row):
            continue
        if i + 1 >= len(raw):
            continue
        next_row = raw.iloc[i + 1].dropna()
        if len(next_row) == 0:
            continue
        next_numeric = sum(1 for v in next_row if parse_numeric(v) is not None)
        score = len(row) + next_numeric * 2
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= 0 else 0


def read_sheet(excel, sheet_name):
    raw = excel.parse(sheet_name, header=None)
    if raw.empty:
        return raw, 0, []
    header_idx = detect_header_row(raw)
    header_vals = raw.iloc[header_idx]
    valid = header_vals.notna()
    if not valid.any():
        data_part = raw.iloc[header_idx + 1 :].copy()
    else:
        data_part = raw.iloc[header_idx + 1 :, valid.values].copy()
        data_part.columns = header_vals[valid].astype(str).values
    excel_rows_full = list(range(header_idx + 2, header_idx + 2 + len(data_part)))
    mask = data_part.notna().any(axis=1).values
    df = data_part[mask].reset_index(drop=True)
    excel_rows = [er for er, keep in zip(excel_rows_full, mask) if keep]
    return df, header_idx, excel_rows


def sanitize_sheet_name(name, used):
    for ch in "[]:*?/\\'":
        name = name.replace(ch, "_")
    name = name.strip()[:31] or "sheet"
    base = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def to_cell_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, float) and val.is_integer():
        return int(val)
    return val


def _display_width(s):
    return sum(2 if ord(c) > 127 else 1 for c in str(s))


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
NUMBER_FORMAT = "#,##0"


def style_header_row(ws, n_cols):
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def auto_size_columns(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        widths = [_display_width(col_name)]
        widths += [_display_width(v) for v in df[col_name].dropna().astype(str)]
        width = min(max(max(widths) + 2, 10), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_dataframe(ws, df, numeric_cols=None):
    numeric_set = set(numeric_cols or [])
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=str(col_name))
    style_header_row(ws, len(df.columns))
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
        for col_idx, val in enumerate(row, 1):
            cell_val = to_cell_value(val)
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell_val is not None:
                cell.value = cell_val
            if df.columns[col_idx - 1] in numeric_set:
                cell.number_format = NUMBER_FORMAT
    auto_size_columns(ws, df)


def render_xlsx_html(xlsx_bytes, sheet_name):
    try:
        src = BytesIO(xlsx_bytes)
        out = StringIO()
        xlsx2html(src, out, sheet=sheet_name)
        return out.getvalue()
    except Exception as e:
        return f"<pre style='color:#c00;padding:8px'>미리보기 렌더링 실패: {e}</pre>"


@st.cache_data(show_spinner=False)
def render_sheet_to_images(xlsx_bytes, sheet_name, dpi=144):
    if not HAS_LIBREOFFICE:
        return []
    try:
        wb = load_workbook(BytesIO(xlsx_bytes))
        if sheet_name not in wb.sheetnames:
            return []
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]
        ws = wb[sheet_name]
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        return []

    with tempfile.TemporaryDirectory() as td:
        xlsx_path = Path(td) / "in.xlsx"
        wb.save(xlsx_path)
        try:
            subprocess.run(
                [LIBREOFFICE_BIN, "--headless", "--convert-to", "pdf",
                 "--outdir", td, str(xlsx_path)],
                check=True, capture_output=True, timeout=60,
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired, FileNotFoundError):
            return []
        pdf_path = Path(td) / "in.pdf"
        if not pdf_path.exists():
            return []
        out_prefix = str(Path(td) / "page")
        try:
            subprocess.run(
                [PDFTOPPM_BIN, "-r", str(dpi), "-png", str(pdf_path), out_prefix],
                check=True, capture_output=True, timeout=60,
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired, FileNotFoundError):
            return []
        return [p.read_bytes() for p in sorted(Path(td).glob("page-*.png"))]


def render_sheet(xlsx_bytes, sheet_name, height=420):
    if HAS_LIBREOFFICE:
        images = render_sheet_to_images(xlsx_bytes, sheet_name)
        if images:
            for img in images:
                st.image(img, width="stretch")
            return
    components.html(
        render_xlsx_html(xlsx_bytes, sheet_name), height=height, scrolling=True
    )


def classify_columns(signature, members):
    numeric = [
        col
        for col in signature
        if all(is_numeric_series(m[2][col]) for m in members)
    ]
    keys = [col for col in signature if col not in numeric]
    return keys, numeric


def union_keys(key_cols, members):
    if not key_cols:
        return None
    parts = [m[2][key_cols] for m in members]
    union = pd.concat(parts, ignore_index=True).drop_duplicates()
    return union.sort_values(key_cols, kind="stable").reset_index(drop=True)


def copy_sheet(src_ws, dst_ws):
    from copy import copy as _copy

    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = col_dim.width
        if col_dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True
    for row_idx, row_dim in src_ws.row_dimensions.items():
        if row_dim.height is not None:
            dst_ws.row_dimensions[row_idx].height = row_dim.height
        if row_dim.hidden:
            dst_ws.row_dimensions[row_idx].hidden = True
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.fill = _copy(cell.fill)
                new_cell.border = _copy(cell.border)
                new_cell.alignment = _copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = _copy(cell.protection)
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))


def _copy_cell_style(src_cell, dst_cell):
    from copy import copy as _copy

    if isinstance(dst_cell, MergedCell):
        return
    if not src_cell.has_style:
        return
    dst_cell.font = _copy(src_cell.font)
    dst_cell.fill = _copy(src_cell.fill)
    dst_cell.border = _copy(src_cell.border)
    dst_cell.alignment = _copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def apply_template_layout(template_ws, dst_ws, header_idx):
    """Copy column widths, pre-header rows, header row (and their merges) from
    template_ws onto dst_ws. Returns the 1-based row index where data begins."""
    for col_letter, col_dim in template_ws.column_dimensions.items():
        if col_dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = col_dim.width

    header_excel_row = header_idx + 1
    for row_idx in range(1, header_excel_row + 1):
        rh = template_ws.row_dimensions.get(row_idx)
        if rh is not None and rh.height is not None:
            dst_ws.row_dimensions[row_idx].height = rh.height
        for col_idx in range(1, template_ws.max_column + 1):
            src_cell = template_ws.cell(row=row_idx, column=col_idx)
            dst_cell = dst_ws.cell(row=row_idx, column=col_idx, value=src_cell.value)
            _copy_cell_style(src_cell, dst_cell)

    for merge_range in template_ws.merged_cells.ranges:
        if merge_range.max_row <= header_excel_row:
            dst_ws.merge_cells(str(merge_range))

    return header_excel_row + 1


def apply_template_row_style(template_ws, template_row, dst_ws, dst_row, n_cols):
    rh = template_ws.row_dimensions.get(template_row)
    if rh is not None and rh.height is not None:
        dst_ws.row_dimensions[dst_row].height = rh.height
    for col_idx in range(1, n_cols + 1):
        _copy_cell_style(
            template_ws.cell(row=template_row, column=col_idx),
            dst_ws.cell(row=dst_row, column=col_idx),
        )


def build_merged(groups, file_bytes_by_name, header_indices):
    wb = Workbook()
    wb.remove(wb.active)
    wb_preview = Workbook()
    wb_preview.remove(wb_preview.active)
    used = set()
    layout = []

    src_wbs = {
        fname: load_workbook(BytesIO(b))
        for fname, b in file_bytes_by_name.items()
    }

    for group_idx, (signature, members) in enumerate(groups.items(), 1):
        if len(members) == 1:
            file_name, sheet_name, _df, _er = members[0]
            stem = file_name.rsplit(".", 1)[0]
            ws_name = sanitize_sheet_name(f"{stem}_{sheet_name}", used)
            src_ws = src_wbs[file_name][sheet_name]
            for target_wb in (wb, wb_preview):
                copy_sheet(src_ws, target_wb.create_sheet(ws_name))
            layout.append((group_idx, None, [ws_name]))
            continue

        key_cols, numeric_cols = classify_columns(signature, members)
        all_keys = union_keys(key_cols, members)
        n_rows = len(all_keys) if key_cols else max(len(m[2]) for m in members)

        source_info = []
        for file_name, sheet_name, df, excel_rows in members:
            stem = file_name.rsplit(".", 1)[0]
            ws_name = sanitize_sheet_name(f"{stem}_{sheet_name}", used)
            src_ws = src_wbs[file_name][sheet_name]
            for target_wb in (wb, wb_preview):
                copy_sheet(src_ws, target_wb.create_sheet(ws_name))
            if key_cols:
                key_map = {}
                for i, row in enumerate(
                    df[key_cols].itertuples(index=False, name=None)
                ):
                    if i < len(excel_rows):
                        key_map.setdefault(row, excel_rows[i])
            else:
                key_map = None
            source_info.append((ws_name, key_map, df, excel_rows))

        # Use first source as styling template for the summary sheet
        template_file, template_sheet, _, template_excel_rows = members[0]
        template_ws = src_wbs[template_file][template_sheet]
        template_header_idx = header_indices.get((template_file, template_sheet), 0)
        template_data_row = template_excel_rows[0] if template_excel_rows else None

        summary_name = sanitize_sheet_name(f"summary_{group_idx}", used)
        for target_wb, formula_mode in ((wb, True), (wb_preview, False)):
            summary = target_wb.create_sheet(summary_name)
            data_start = apply_template_layout(template_ws, summary, template_header_idx)

            for row_idx in range(n_rows):
                dst_row = data_start + row_idx
                if template_data_row is not None:
                    apply_template_row_style(
                        template_ws, template_data_row, summary, dst_row, len(signature)
                    )
                key_tuple = (
                    tuple(all_keys.iloc[row_idx][key_cols]) if key_cols else None
                )
                for col_idx, col_name in enumerate(signature, 1):
                    col_letter = get_column_letter(col_idx)
                    cell = summary.cell(row=dst_row, column=col_idx)
                    if col_name in numeric_cols:
                        refs = []
                        values = []
                        for ws_name, key_map, df_src, _ in source_info:
                            if key_cols and key_map and key_tuple in key_map:
                                src_row = key_map[key_tuple]
                                refs.append(f"'{ws_name}'!{col_letter}{src_row}")
                                match = df_src[
                                    df_src[key_cols]
                                    .apply(tuple, axis=1) == key_tuple
                                ]
                                if not match.empty:
                                    v = parse_numeric(match.iloc[0][col_name])
                                    if v is not None:
                                        values.append(v)
                        if formula_mode:
                            cell.value = f"=SUM({','.join(refs)})" if refs else 0
                        else:
                            cell.value = to_cell_value(sum(values)) if values else 0
                        if not cell.number_format or cell.number_format == "General":
                            cell.number_format = NUMBER_FORMAT
                    elif key_cols:
                        val = all_keys.iloc[row_idx][col_name]
                        cell.value = to_cell_value(val)

        n_sources = len(source_info)
        wb.move_sheet(summary_name, offset=-n_sources)
        wb_preview.move_sheet(summary_name, offset=-n_sources)

        layout.append(
            (group_idx, summary_name, [info[0] for info in source_info])
        )

    return wb, wb_preview, layout


def _get_merged_value(ws, row_idx, col_idx):
    """Value at (row, col), looking up the merged-region anchor if cell itself is empty."""
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        return cell.value
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row_idx <= mr.max_row and mr.min_col <= col_idx <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def _build_combined_headers(ws, header_rows_1based, n_cols):
    """Combine values from multiple header rows (1-based) into one name per column.
    Merged super-headers are unfolded via merge-anchor lookup."""
    per_row = [
        [_get_merged_value(ws, r, c) for c in range(1, n_cols + 1)]
        for r in header_rows_1based
    ]
    headers = []
    for c_idx in range(n_cols):
        parts = []
        for row_vals in per_row:
            v = row_vals[c_idx]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if parts and parts[-1] == s:
                continue  # avoid duplicating identical labels across rows
            parts.append(s)
        headers.append(" / ".join(parts) if parts else None)
    return headers


def parse_sheet_with_overrides(
    xlsx_bytes,
    sheet_name,
    header_row=None,
    header_rows=None,
    skip_rows=(),
    skip_keywords=(),
    forward_fill_columns=(),
):
    """Read a sheet honoring multi-row headers, explicit skip rows, keyword-based
    row skips, and forward-fill on sparse category columns.

    Returns (df, excel_rows, original_cols) where excel_rows[i] is the 1-based
    source row for df.iloc[i] and original_cols[i] is the 1-based source column
    for df.columns[i]."""
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(), [], []
    ws = wb[sheet_name]
    n_cols = ws.max_column or 0
    n_rows_total = ws.max_row or 0
    if n_cols == 0 or n_rows_total == 0:
        return pd.DataFrame(), [], []

    if header_rows:
        h_rows = sorted(int(r) for r in header_rows)
    elif header_row is not None:
        h_rows = [int(header_row)]
    else:
        h_rows = [0]
    last_header_row = h_rows[-1]
    headers = _build_combined_headers(ws, [r + 1 for r in h_rows], n_cols)

    valid_cols = [i for i, h in enumerate(headers) if h]
    if not valid_cols:
        return pd.DataFrame(), [], []
    raw_final_headers = [headers[i] for i in valid_cols]

    # De-duplicate column names. When LLM picks header_rows that doesn't include
    # the super-header row, multiple columns can end up named "합계" — pandas then
    # refuses many operations. Suffix duplicates so labels are unique.
    counts = {}
    final_headers = []
    for h in raw_final_headers:
        if h not in counts:
            counts[h] = 1
            final_headers.append(h)
        else:
            counts[h] += 1
            final_headers.append(f"{h}_{counts[h]}")
    original_cols = [i + 1 for i in valid_cols]

    data_first_excel_row = last_header_row + 2
    data_rows_raw = []
    excel_rows_full = []
    for r in range(data_first_excel_row, n_rows_total + 1):
        row_vals = [_get_merged_value(ws, r, c) for c in range(1, n_cols + 1)]
        data_rows_raw.append([row_vals[i] for i in valid_cols])
        excel_rows_full.append(r)

    df = pd.DataFrame(data_rows_raw, columns=final_headers)

    skip_set = set(skip_rows or [])
    keep_mask = [er - 1 not in skip_set for er in excel_rows_full]
    if skip_keywords:
        cleaned_keywords = [str(k).strip() for k in skip_keywords if k and str(k).strip()]
        for i, row in enumerate(data_rows_raw):
            if not keep_mask[i] or not row:
                continue
            first = row[0]
            if first is None:
                continue
            s = str(first).strip()
            if any(kw in s for kw in cleaned_keywords):
                keep_mask[i] = False
    notna_mask = df.notna().any(axis=1).tolist()
    final_mask = [k and n for k, n in zip(keep_mask, notna_mask)]

    df = df[final_mask].reset_index(drop=True)
    excel_rows = [er for er, keep in zip(excel_rows_full, final_mask) if keep]

    if forward_fill_columns:
        for col in forward_fill_columns:
            if col in df.columns:
                df[col] = df[col].ffill()

    return df, excel_rows, original_cols


AGG_FORMULA = {
    "sum": "SUM", "mean": "AVERAGE", "average": "AVERAGE", "avg": "AVERAGE",
    "max": "MAX", "min": "MIN",
}


def _agg_values(values, agg):
    if not values:
        return 0
    agg = (agg or "sum").lower()
    if agg == "sum":
        return sum(values)
    if agg in ("mean", "average", "avg"):
        return sum(values) / len(values)
    if agg == "max":
        return max(values)
    if agg == "min":
        return min(values)
    if agg == "latest":
        return values[-1]
    return sum(values)


def _read_sheet_for_union(ws, header_n, key_n, skip_keywords):
    """Return ordered list of (excel_row, key_tuple) for data rows.
    Skips fully-empty rows and rows whose first key cell contains a skip keyword."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    cleaned_kw = [str(k).strip() for k in (skip_keywords or []) if k and str(k).strip()]
    data_rows = []
    for r in range(header_n + 1, max_row + 1):
        row_vals = [_get_merged_value(ws, r, c) for c in range(1, max_col + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue
        key = tuple(
            (str(v).strip() if v is not None else None) for v in row_vals[:key_n]
        )
        first_key = next((k for k in key if k), None)
        if cleaned_kw and first_key and any(kw in first_key for kw in cleaned_kw):
            continue
        data_rows.append((r, key))
    return data_rows


def _parse_int(value, default):
    try:
        if isinstance(value, list) and value:
            return max(int(v) for v in value)
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalize_sheet_spec(spec):
    """Coerce a plan sheet spec into (header_n, key_n, skip_keywords)."""
    header_n = spec.get("header_row_count")
    if header_n is None:
        if isinstance(spec.get("header_rows"), list) and spec.get("header_rows"):
            header_n = max(int(r) for r in spec["header_rows"]) + 1
        elif "header_row" in spec:
            try:
                header_n = int(spec["header_row"]) + 1
            except (TypeError, ValueError):
                header_n = 1
        else:
            header_n = 1
    try:
        header_n = int(header_n)
    except (TypeError, ValueError):
        header_n = 1
    if header_n < 1:
        header_n = 1
    try:
        key_n = int(spec.get("key_column_count") or 1)
    except (TypeError, ValueError):
        key_n = 1
    if key_n < 1:
        key_n = 1
    skip_kw = spec.get("skip_keywords") or []
    return header_n, key_n, skip_kw


def _build_summary_into(target_wb, tpl_sheet_name, members, group_idx, formula_mode):
    """Add a summary sheet to `target_wb`. Returns the summary sheet name."""
    if tpl_sheet_name not in target_wb.sheetnames:
        return None
    template_ws = target_wb[tpl_sheet_name]
    summary_ws = target_wb.copy_worksheet(template_ws)
    summary_name = f"summary_{group_idx}"
    # copy_worksheet picks its own title; rename
    existing = set(target_wb.sheetnames)
    n = 2
    final = summary_name
    while final in existing and final != summary_ws.title:
        final = f"{summary_name}_{n}"
        n += 1
    summary_ws.title = final

    tpl = members[0]
    tpl_max_col = tpl["max_col"]
    tpl_key_n = tpl["key_n"]
    tpl_header_n = tpl["header_n"]
    tpl_data_row_for_style = tpl["data_rows"][0][0] if tpl["data_rows"] else (
        tpl_header_n + 1
    )

    def collect_refs_and_values(key, col):
        refs, values = [], []
        for m in members:
            keys_in_m = {k for _, k in m["data_rows"]}
            if key not in keys_in_m:
                continue
            src_row = next(r for r, k in m["data_rows"] if k == key)
            if col > m["max_col"]:
                continue
            refs.append(f"'{m['merged_sheet_name']}'!{get_column_letter(col)}{src_row}")
            v = m["ws"].cell(row=src_row, column=col).value
            vn = parse_numeric(v)
            if vn is not None:
                values.append(vn)
        return refs, values

    # 1) Replace measure cells in template's data rows
    for excel_row, key in tpl["data_rows"]:
        for col in range(tpl_key_n + 1, tpl_max_col + 1):
            refs, values = collect_refs_and_values(key, col)
            if not refs:
                continue
            cell = summary_ws.cell(row=excel_row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if formula_mode:
                cell.value = f"=SUM({','.join(refs)})"
            else:
                cell.value = to_cell_value(sum(values)) if values else 0

    # 2) Append new keys at bottom
    tpl_key_set = {k for _, k in tpl["data_rows"]}
    appended = set()
    next_row = (summary_ws.max_row or tpl_header_n) + 1
    for m in members[1:]:
        for _src_row, key in m["data_rows"]:
            if key in tpl_key_set or key in appended:
                continue
            appended.add(key)
            apply_template_row_style(
                tpl["ws"], tpl_data_row_for_style,
                summary_ws, next_row, tpl_max_col,
            )
            for ci in range(tpl_key_n):
                v = key[ci] if ci < len(key) else None
                if v is None:
                    continue
                cell = summary_ws.cell(row=next_row, column=ci + 1)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = v
            for col in range(tpl_key_n + 1, tpl_max_col + 1):
                refs, values = collect_refs_and_values(key, col)
                if not refs:
                    continue
                cell = summary_ws.cell(row=next_row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                if formula_mode:
                    cell.value = f"=SUM({','.join(refs)})"
                else:
                    cell.value = to_cell_value(sum(values)) if values else 0
            next_row += 1

    return final


def build_from_plan(plan, file_bytes_by_name):
    """Hybrid build:
    - Phase 1 (XML-level): merge all source files into one xlsx with byte-faithful
      source sheets, unified styles/sharedStrings tables.
    - Phase 2 (openpyxl): clone the template sheet per group as a summary sheet,
      then replace measure cells with =SUM formulas referencing the merged-in
      source sheets by their final names."""
    # Collect unique source files referenced by the plan (preserve order)
    used_files: list[str] = []
    seen_files: set[str] = set()
    for g in plan.get("groups", []) or []:
        for s in g.get("sheets", []) or []:
            fn = s.get("file")
            if fn and fn in file_bytes_by_name and fn not in seen_files:
                seen_files.add(fn)
                used_files.append(fn)

    if not used_files:
        wb = Workbook(); wb.remove(wb.active)
        return wb, Workbook(), []

    sources = [(fn, file_bytes_by_name[fn]) for fn in used_files]
    merged_bytes, name_map = xml_merge_workbooks(sources)

    wb_download = load_workbook(BytesIO(merged_bytes))
    wb_preview = load_workbook(BytesIO(merged_bytes))

    layout = []
    for group_idx, group in enumerate(plan.get("groups", []), 1):
        sheet_specs = group.get("sheets", []) or []
        if not sheet_specs:
            continue

        # Build per-member metadata (against the DOWNLOAD wb; preview shares names)
        members_dl, members_pv = [], []
        members_sheet_names = []
        for spec in sheet_specs:
            fn = spec.get("file"); sn = spec.get("sheet")
            if fn not in file_bytes_by_name:
                continue
            merged_name = name_map.get((fn, sn))
            if not merged_name or merged_name not in wb_download.sheetnames:
                continue
            header_n, key_n, skip_kw = _normalize_sheet_spec(spec)
            ws_dl = wb_download[merged_name]
            ws_pv = wb_preview[merged_name]
            data_rows = _read_sheet_for_union(ws_dl, header_n, key_n, skip_kw)
            common = {
                "header_n": header_n, "key_n": key_n, "skip_kw": skip_kw,
                "data_rows": data_rows,
                "max_col": ws_dl.max_column or 0,
                "merged_sheet_name": merged_name,
            }
            members_dl.append({**common, "ws": ws_dl})
            members_pv.append({**common, "ws": ws_pv})
            members_sheet_names.append(merged_name)

        if not members_dl:
            continue

        if len(members_dl) == 1:
            layout.append((group_idx, None, [members_sheet_names[0]]))
            continue

        tpl_name = members_sheet_names[0]
        summary_name_dl = _build_summary_into(
            wb_download, tpl_name, members_dl, group_idx, formula_mode=True
        )
        summary_name_pv = _build_summary_into(
            wb_preview, tpl_name, members_pv, group_idx, formula_mode=False
        )

        if summary_name_dl:
            try:
                idx = wb_download.sheetnames.index(tpl_name)
                cur = wb_download.sheetnames.index(summary_name_dl)
                wb_download.move_sheet(summary_name_dl, offset=idx - cur)
            except ValueError:
                pass
        if summary_name_pv:
            try:
                idx = wb_preview.sheetnames.index(tpl_name)
                cur = wb_preview.sheetnames.index(summary_name_pv)
                wb_preview.move_sheet(summary_name_pv, offset=idx - cur)
            except ValueError:
                pass

        layout.append((group_idx, summary_name_dl, members_sheet_names))

    return wb_download, wb_preview, layout


with st.sidebar:
    st.header("엑셀 업로드")
    uploaded_files = st.file_uploader(
        "통합할 엑셀 파일을 선택하세요",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

if not uploaded_files:
    st.info("왼쪽 사이드바에서 엑셀 파일을 업로드하세요.")
    st.stop()

all_sheets = []
header_indices = {}
file_bytes_by_name = {}
for file in uploaded_files:
    raw_bytes = file.getvalue()
    file_bytes_by_name[file.name] = raw_bytes
    try:
        excel = pd.ExcelFile(BytesIO(raw_bytes))
    except Exception as e:
        st.error(f"{file.name}: 파일을 읽지 못했습니다 ({e})")
        continue
    for sheet_name in excel.sheet_names:
        df, header_idx, excel_rows = read_sheet(excel, sheet_name)
        all_sheets.append((file.name, sheet_name, df, excel_rows))
        header_indices[(file.name, sheet_name)] = header_idx

if not all_sheets:
    st.stop()

with st.expander("원본 미리보기", expanded=False):
    if not HAS_LIBREOFFICE:
        st.caption(
            "원본 정확 렌더링을 위해 LibreOffice + poppler-utils 설치를 권장합니다. "
            "(`sudo apt install -y libreoffice-calc poppler-utils`) — 현재는 HTML 폴백으로 표시됩니다."
        )
    file_names = list(dict.fromkeys(f for f, _, _, _ in all_sheets))
    for file_name in file_names:
        st.markdown(f"**{file_name}**")
        members = [(s, df) for f, s, df, _ in all_sheets if f == file_name]
        tabs = st.tabs([s for s, _ in members])
        for tab, (sheet_name, df) in zip(tabs, members):
            with tab:
                h_idx = header_indices.get((file_name, sheet_name), 0)
                st.caption(
                    f"{len(df)} rows × {len(df.columns)} cols · 감지된 헤더 행: {h_idx} (0-based)"
                )
                render_sheet(file_bytes_by_name[file_name], sheet_name)

groups = defaultdict(list)
for file_name, sheet_name, df, excel_rows in all_sheets:
    groups[tuple(df.columns)].append((file_name, sheet_name, df, excel_rows))

st.header("통합 엑셀")
st.caption(
    "같은 컬럼 구조 시트끼리 묶어, 텍스트 컬럼은 키 / 숫자 컬럼은 통합 대상으로 분류합니다. "
    "합산 시트(SUM 수식)가 그룹 앞에 들어가고 원본 시트는 키 기준 정렬되어 뒤따릅니다."
)

for group_idx, (signature, members) in enumerate(groups.items(), 1):
    key_cols, numeric_cols = classify_columns(signature, members)
    sources = [f"{f} / {s}" for f, s, _, _ in members]
    st.markdown(
        f"**그룹 {group_idx}** · {len(members)}시트 — {' · '.join(sources)}"
    )
    st.caption(
        f"키: {', '.join(key_cols) or '(없음)'}  ·  숫자: {', '.join(numeric_cols) or '(없음)'}"
    )

st.write("")

col_btn, col_dl = st.columns([1, 1])
with col_btn:
    if st.button("통합 엑셀 생성", type="primary", width="stretch"):
        wb, wb_preview, layout = build_merged(groups, file_bytes_by_name, header_indices)
        buf = BytesIO(); wb.save(buf)
        buf_p = BytesIO(); wb_preview.save(buf_p)
        st.session_state["merged_xlsx"] = buf.getvalue()
        st.session_state["merged_preview_xlsx"] = buf_p.getvalue()
        st.session_state["merged_layout"] = layout
with col_dl:
    if "merged_xlsx" in st.session_state:
        st.download_button(
            "merged.xlsx 다운로드",
            data=st.session_state["merged_xlsx"],
            file_name="merged.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

if "merged_layout" in st.session_state and "merged_preview_xlsx" in st.session_state:
    st.divider()
    st.subheader("결과 미리보기")
    st.caption("실제 파일의 합산 시트에는 SUM 수식이 들어가며, 미리보기에는 계산값으로 표시됩니다.")
    preview_bytes = st.session_state["merged_preview_xlsx"]
    for group_idx, summary_name, source_names in st.session_state["merged_layout"]:
        with st.expander(f"그룹 {group_idx}", expanded=True):
            if summary_name is not None:
                labels = [f"{summary_name} (합산)"] + source_names
                tabs = st.tabs(labels)
                with tabs[0]:
                    render_sheet(preview_bytes, summary_name)
                for tab, name in zip(tabs[1:], source_names):
                    with tab:
                        render_sheet(preview_bytes, name)
            else:
                name = source_names[0]
                st.caption(name)
                render_sheet(preview_bytes, name)

st.divider()
st.header("AI 분석 & 챗봇")
st.caption(
    f"LLM이 시트 구조(헤더/소계/키/측정값)를 분석해 통합 plan을 만들고, "
    f"챗봇으로 plan을 다듬은 뒤 실행합니다. 모델: `{OLLAMA_MODEL}`"
)

if "ai_notes" not in st.session_state:
    st.session_state["ai_notes"] = {}
if "ai_plan" not in st.session_state:
    st.session_state["ai_plan"] = None
if "ai_chat" not in st.session_state:
    st.session_state["ai_chat"] = []

col_a, col_b = st.columns([1, 1])
with col_a:
    if st.button(
        "1. 파일별 분석 노트 생성", type="secondary", width="stretch",
        disabled=not uploaded_files,
    ):
        with st.spinner("각 파일을 분석 중... (파일당 5-15초)"):
            notes = {}
            try:
                for fn in file_bytes_by_name:
                    sense = sense_workbook(file_bytes_by_name[fn], fn)
                    notes[fn] = analyze_per_file(sense)
                st.session_state["ai_notes"] = notes
                st.session_state["ai_plan"] = None
                st.session_state["ai_chat"] = []
            except Exception as e:
                st.error(f"노트 생성 실패: {e}")
with col_b:
    if st.button("plan/대화 초기화", width="stretch"):
        st.session_state["ai_notes"] = {}
        st.session_state["ai_plan"] = None
        st.session_state["ai_chat"] = []
        for k in (
            "ai_merged_xlsx", "ai_merged_preview_xlsx", "ai_merged_layout",
            "ai_plan_error", "ai_raw_response", "ai_plan_elapsed",
            "ai_plan_fallback",
            "ai_refine_error", "ai_refine_raw", "ai_refine_elapsed",
            "ai_integrate_error",
        ):
            st.session_state.pop(k, None)

notes_state = st.session_state["ai_notes"]
if notes_state:
    st.subheader("파일별 분석 노트")
    st.caption("LLM 초안입니다. 잘못된 부분은 직접 고쳐주세요. plan 생성에 그대로 반영됩니다.")
    edited = {}
    for fn, note in notes_state.items():
        edited[fn] = st.text_area(
            fn, value=note, height=200, key=f"note_{fn}",
        )
    st.session_state["ai_notes"] = edited

    if st.button("2. plan 생성", type="primary", width="stretch"):
        import time, traceback
        st.session_state["ai_plan_error"] = None
        st.session_state["ai_raw_response"] = None
        st.session_state["ai_plan_elapsed"] = None
        st.session_state["ai_plan_fallback"] = False
        with st.spinner("plan 생성 중... (실패 시 LLM이 자가 수정으로 재시도)"):
            t0 = time.time()
            try:
                sense_blocks = [
                    sense_workbook(file_bytes_by_name[fn], fn)
                    for fn in file_bytes_by_name
                ]
                notes_section = "\n\n".join(
                    f"--- {fn}에 대한 분석 노트 ---\n{(note or '').strip()}"
                    for fn, note in edited.items() if note and note.strip()
                )
                parts = []
                if notes_section:
                    parts.append("각 파일에 대한 사전 분석 노트:\n\n" + notes_section)
                parts.append("Excel 시트 원본 스냅샷:\n\n" + "\n\n".join(sense_blocks))
                parts.append("위 노트와 스냅샷을 바탕으로 통합 plan JSON을 출력해주세요.")
                msgs = [
                    {"role": "system", "content": ANALYZE_SYSTEM},
                    {"role": "user", "content": "\n\n".join(parts)},
                ]
                parsed, raw, err, attempts = llm_get_json_with_retry(msgs, max_attempts=3)
                st.session_state["ai_raw_response"] = raw
                if parsed:
                    st.session_state["ai_plan"] = parsed
                    st.session_state["ai_chat"] = []
                    if len(attempts) > 1:
                        st.session_state["ai_plan_error"] = (
                            f"LLM이 {len(attempts)}번 시도 후 성공했어요."
                        )
                else:
                    st.session_state["ai_plan"] = build_heuristic_plan(file_bytes_by_name)
                    st.session_state["ai_plan_fallback"] = True
                    st.session_state["ai_chat"] = []
                    st.session_state["ai_plan_error"] = (
                        f"LLM이 {len(attempts)}번 시도했지만 유효한 JSON을 못 만들어서 "
                        f"휴리스틱 plan으로 대체했어요. 챗봇으로 보정하거나 다시 시도해주세요. "
                        f"마지막 오류: {err}"
                    )
            except Exception as e:
                st.session_state["ai_plan_error"] = (
                    f"plan 생성 실패: {type(e).__name__}: {e}\n\n"
                    f"```\n{traceback.format_exc()}\n```"
                )
                try:
                    st.session_state["ai_plan"] = build_heuristic_plan(file_bytes_by_name)
                    st.session_state["ai_plan_fallback"] = True
                    st.session_state["ai_chat"] = []
                except Exception:
                    pass
            finally:
                st.session_state["ai_plan_elapsed"] = time.time() - t0

if st.session_state.get("ai_plan_error"):
    st.error(st.session_state["ai_plan_error"])
if st.session_state.get("ai_plan_elapsed"):
    st.caption(
        f"마지막 plan 호출 응답 시간: {st.session_state['ai_plan_elapsed']:.1f}초"
    )
if st.session_state.get("ai_raw_response"):
    with st.expander("LLM 원시 응답 (debug)", expanded=False):
        st.code(st.session_state["ai_raw_response"], language="json")

plan = st.session_state.get("ai_plan")
if plan:
    if st.session_state.get("ai_plan_fallback"):
        st.warning(
            "현재 plan은 LLM 실패로 인한 휴리스틱 fallback입니다. "
            "챗봇으로 의도를 알려주거나 '2. plan 생성'을 다시 눌러주세요."
        )
    with st.expander("현재 plan (JSON)", expanded=False):
        st.json(plan)

    groups = plan.get("groups") or []
    if groups:
        st.markdown("**plan 요약**")
        for i, g in enumerate(groups, 1):
            label = g.get("label") or f"그룹 {i}"
            sheets = g.get("sheets") or []
            st.markdown(f"- **{label}** · {len(sheets)}개 시트")
            for s in sheets:
                hn = s.get("header_row_count") or s.get("header_rows") or "?"
                kn = s.get("key_column_count") or "?"
                skip = s.get("skip_keywords") or []
                st.caption(
                    f"  · {s.get('file','')}/{s.get('sheet','')} "
                    f"· 헤더 {hn}행 · 키 {kn}열"
                    + (f" · skip: {', '.join(skip)}" if skip else "")
                )

    ambs = plan.get("ambiguities") or []
    if ambs:
        st.info("LLM 질문:\n" + "\n".join(f"- {q}" for q in ambs))

    for msg in st.session_state["ai_chat"]:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    user_msg = st.chat_input("plan을 수정하거나 질문에 답하세요...")
    if user_msg:
        import time, traceback
        st.session_state["ai_chat"].append({"role": "user", "content": user_msg})
        st.session_state["ai_refine_error"] = None
        st.session_state["ai_refine_raw"] = None
        st.session_state["ai_refine_elapsed"] = None
        with st.spinner("LLM이 plan을 갱신 중... (실패 시 자가 수정으로 재시도)"):
            t0 = time.time()
            try:
                msgs = [{"role": "system", "content": REFINE_SYSTEM}]
                msgs.append({
                    "role": "user",
                    "content": f"현재 plan:\n{json.dumps(plan, ensure_ascii=False, indent=2)}",
                })
                msgs.append({"role": "assistant", "content": "확인했습니다."})
                for turn in st.session_state["ai_chat"][-6:]:
                    msgs.append({"role": turn["role"], "content": turn["content"]})
                parsed, raw, err, attempts = llm_get_json_with_retry(msgs, max_attempts=3)
                st.session_state["ai_refine_raw"] = raw
                if parsed:
                    st.session_state["ai_plan"] = parsed
                    info = (
                        "plan을 업데이트했어요."
                        if len(attempts) <= 1
                        else f"plan을 업데이트했어요. ({len(attempts)}회 재시도 후 성공)"
                    )
                    st.session_state["ai_chat"].append({
                        "role": "assistant", "content": info,
                    })
                else:
                    st.session_state["ai_refine_error"] = (
                        f"LLM이 {len(attempts)}번 시도했지만 유효한 JSON을 못 만들었어요 — "
                        f"plan은 그대로 유지합니다. 마지막 오류: {err}"
                    )
                    st.session_state["ai_chat"].append({
                        "role": "assistant",
                        "content": (
                            f"갱신 실패 (LLM이 JSON을 못 만듦). 다시 표현을 바꿔 시도해주세요."
                        ),
                    })
            except Exception as e:
                tb = traceback.format_exc()
                st.session_state["ai_refine_error"] = (
                    f"{type(e).__name__}: {e}\n```\n{tb}\n```"
                )
                st.session_state["ai_chat"].append({
                    "role": "assistant", "content": f"갱신 실패: {e}",
                })
            finally:
                st.session_state["ai_refine_elapsed"] = time.time() - t0
        st.rerun()

    if st.session_state.get("ai_refine_error"):
        st.error(st.session_state["ai_refine_error"])
    if st.session_state.get("ai_refine_elapsed"):
        st.caption(
            f"마지막 refine 응답 시간: {st.session_state['ai_refine_elapsed']:.1f}초"
        )
    if st.session_state.get("ai_refine_raw"):
        with st.expander("Refine LLM 원시 응답 (debug)", expanded=False):
            st.code(st.session_state["ai_refine_raw"], language="json")

    st.divider()
    if st.button("AI plan으로 통합 실행", type="primary", width="stretch"):
        import traceback
        st.session_state["ai_integrate_error"] = None
        with st.spinner("plan에 따라 통합 중..."):
            try:
                wb_ai, wb_ai_preview, layout_ai = build_from_plan(plan, file_bytes_by_name)
                buf = BytesIO(); wb_ai.save(buf)
                buf_p = BytesIO(); wb_ai_preview.save(buf_p)
                st.session_state["ai_merged_xlsx"] = buf.getvalue()
                st.session_state["ai_merged_preview_xlsx"] = buf_p.getvalue()
                st.session_state["ai_merged_layout"] = layout_ai
            except Exception as e:
                tb = traceback.format_exc()
                st.session_state["ai_integrate_error"] = (
                    f"{type(e).__name__}: {e}\n```\n{tb}\n```"
                )

    if st.session_state.get("ai_integrate_error"):
        st.error(st.session_state["ai_integrate_error"])

if "ai_merged_xlsx" in st.session_state:
    st.download_button(
        "merged_ai.xlsx 다운로드",
        data=st.session_state["ai_merged_xlsx"],
        file_name="merged_ai.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )

if "ai_merged_layout" in st.session_state and "ai_merged_preview_xlsx" in st.session_state:
    st.subheader("AI 통합 결과 미리보기")
    preview_bytes_ai = st.session_state["ai_merged_preview_xlsx"]
    for group_idx, summary_name, source_names in st.session_state["ai_merged_layout"]:
        with st.expander(f"그룹 {group_idx}", expanded=True):
            if summary_name is not None:
                labels = [f"{summary_name} (합산)"] + source_names
                tabs = st.tabs(labels)
                with tabs[0]:
                    render_sheet(preview_bytes_ai, summary_name)
                for tab, name in zip(tabs[1:], source_names):
                    with tab:
                        render_sheet(preview_bytes_ai, name)
            else:
                name = source_names[0]
                st.caption(name)
                render_sheet(preview_bytes_ai, name)
